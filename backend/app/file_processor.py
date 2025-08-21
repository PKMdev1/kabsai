import os
import magic
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import json
import csv
import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Optional, Tuple
import tiktoken
import re

class FileProcessor:
    def __init__(self):
        self.supported_formats = {
            'application/pdf': self._extract_pdf_text,
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': self._extract_docx_text,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': self._extract_xlsx_text,
            'text/plain': self._extract_txt_text,
            'text/csv': self._extract_csv_text,
            'application/json': self._extract_json_text,
            'text/markdown': self._extract_md_text,
            'text/html': self._extract_html_text,
            'application/xml': self._extract_xml_text,
            'text/xml': self._extract_xml_text,
        }
        
        # Initialize tokenizer
        self.tokenizer = tiktoken.get_encoding("cl100k_base")
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get basic file information."""
        file_size = os.path.getsize(file_path)
        mime_type = magic.from_file(file_path, mime=True)
        
        return {
            'size': file_size,
            'mime_type': mime_type,
            'extension': os.path.splitext(file_path)[1].lower(),
            'filename': os.path.basename(file_path)
        }
    
    def extract_text(self, file_path: str) -> str:
        """Extract text content from file based on its type."""
        mime_type = magic.from_file(file_path, mime=True)
        
        if mime_type in self.supported_formats:
            return self.supported_formats[mime_type](file_path)
        else:
            raise ValueError(f"Unsupported file type: {mime_type}")
    
    def extract_xml_data(self, file_path: str) -> Dict[str, Any]:
        """Extract structured data from XML file."""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Extract basic XML structure
            xml_data = {
                'root_tag': root.tag,
                'attributes': dict(root.attrib),
                'elements': self._parse_xml_elements(root),
                'text_content': self._extract_xml_text(file_path),
                'structure': self._get_xml_structure(root)
            }
            
            return xml_data
        except ET.ParseError as e:
            raise ValueError(f"Invalid XML file: {e}")
        except Exception as e:
            raise ValueError(f"Error processing XML file: {e}")
    
    def match_xml_data(self, xml_data: Dict[str, Any], search_criteria: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Match XML data based on search criteria."""
        matches = []
        
        def search_elements(elements: List[Dict[str, Any]], criteria: Dict[str, Any]) -> List[Dict[str, Any]]:
            results = []
            
            for element in elements:
                match = True
                
                # Check tag name
                if 'tag' in criteria and criteria['tag']:
                    if not re.search(criteria['tag'], element.get('tag', ''), re.IGNORECASE):
                        match = False
                
                # Check attributes
                if 'attributes' in criteria and criteria['attributes']:
                    element_attrs = element.get('attributes', {})
                    for attr_key, attr_value in criteria['attributes'].items():
                        if attr_key not in element_attrs or not re.search(attr_value, str(element_attrs[attr_key]), re.IGNORECASE):
                            match = False
                            break
                
                # Check text content
                if 'text' in criteria and criteria['text']:
                    element_text = element.get('text', '')
                    if not re.search(criteria['text'], element_text, re.IGNORECASE):
                        match = False
                
                # Check value
                if 'value' in criteria and criteria['value']:
                    element_value = element.get('value', '')
                    if not re.search(criteria['value'], str(element_value), re.IGNORECASE):
                        match = False
                
                if match:
                    results.append(element)
                
                # Recursively search child elements
                if 'children' in element:
                    child_results = search_elements(element['children'], criteria)
                    results.extend(child_results)
            
            return results
        
        # Search in elements
        if 'elements' in xml_data:
            element_matches = search_elements(xml_data['elements'], search_criteria)
            matches.extend(element_matches)
        
        return matches
    
    def extract_xml_schema(self, file_path: str) -> Dict[str, Any]:
        """Extract XML schema information."""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            schema = {
                'root_element': root.tag,
                'namespaces': self._extract_namespaces(root),
                'element_hierarchy': self._build_element_hierarchy(root),
                'attribute_definitions': self._extract_attribute_definitions(root),
                'data_types': self._infer_data_types(root)
            }
            
            return schema
        except Exception as e:
            raise ValueError(f"Error extracting XML schema: {e}")
    
    def _parse_xml_elements(self, element: ET.Element) -> List[Dict[str, Any]]:
        """Parse XML elements recursively."""
        elements = []
        
        for child in element:
            child_data = {
                'tag': child.tag,
                'attributes': dict(child.attrib),
                'text': child.text.strip() if child.text and child.text.strip() else '',
                'value': child.text.strip() if child.text and child.text.strip() else '',
                'children': self._parse_xml_elements(child)
            }
            elements.append(child_data)
        
        return elements
    
    def _get_xml_structure(self, element: ET.Element, max_depth: int = 5) -> Dict[str, Any]:
        """Get XML structure information."""
        def build_structure(elem: ET.Element, depth: int = 0) -> Dict[str, Any]:
            if depth > max_depth:
                return {'type': 'truncated'}
            
            structure = {
                'tag': elem.tag,
                'attributes': list(elem.attrib.keys()),
                'has_text': bool(elem.text and elem.text.strip()),
                'children': []
            }
            
            for child in elem:
                child_structure = build_structure(child, depth + 1)
                structure['children'].append(child_structure)
            
            return structure
        
        return build_structure(element)
    
    def _extract_namespaces(self, element: ET.Element) -> Dict[str, str]:
        """Extract XML namespaces."""
        namespaces = {}
        
        # Extract from root element
        for key, value in element.attrib.items():
            if key.startswith('xmlns:'):
                prefix = key.split(':', 1)[1]
                namespaces[prefix] = value
            elif key == 'xmlns':
                namespaces['default'] = value
        
        return namespaces
    
    def _build_element_hierarchy(self, element: ET.Element) -> Dict[str, Any]:
        """Build element hierarchy."""
        hierarchy = {
            'tag': element.tag,
            'children': {}
        }
        
        for child in element:
            if child.tag not in hierarchy['children']:
                hierarchy['children'][child.tag] = {
                    'count': 0,
                    'attributes': set(),
                    'has_text': False
                }
            
            hierarchy['children'][child.tag]['count'] += 1
            hierarchy['children'][child.tag]['attributes'].update(child.attrib.keys())
            if child.text and child.text.strip():
                hierarchy['children'][child.tag]['has_text'] = True
        
        # Convert sets to lists for JSON serialization
        for child_info in hierarchy['children'].values():
            child_info['attributes'] = list(child_info['attributes'])
        
        return hierarchy
    
    def _extract_attribute_definitions(self, element: ET.Element) -> Dict[str, List[str]]:
        """Extract attribute definitions from XML."""
        attributes = {}
        
        def collect_attributes(elem: ET.Element):
            for attr_name, attr_value in elem.attrib.items():
                if attr_name not in attributes:
                    attributes[attr_name] = []
                if attr_value not in attributes[attr_name]:
                    attributes[attr_name].append(attr_value)
            
            for child in elem:
                collect_attributes(child)
        
        collect_attributes(element)
        return attributes
    
    def _infer_data_types(self, element: ET.Element) -> Dict[str, str]:
        """Infer data types from XML content."""
        data_types = {}
        
        def analyze_element(elem: ET.Element):
            if elem.text and elem.text.strip():
                text = elem.text.strip()
                
                # Try to infer data type
                if text.isdigit():
                    data_types[elem.tag] = 'integer'
                elif text.replace('.', '').replace('-', '').isdigit():
                    data_types[elem.tag] = 'float'
                elif text.lower() in ['true', 'false']:
                    data_types[elem.tag] = 'boolean'
                elif re.match(r'\d{4}-\d{2}-\d{2}', text):
                    data_types[elem.tag] = 'date'
                elif re.match(r'\d{2}:\d{2}:\d{2}', text):
                    data_types[elem.tag] = 'time'
                else:
                    data_types[elem.tag] = 'string'
            
            for child in elem:
                analyze_element(child)
        
        analyze_element(element)
        return data_types
    
    def _extract_xml_text(self, file_path: str) -> str:
        """Extract text content from XML file."""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Extract all text content
            text_parts = []
            
            def extract_text(element):
                if element.text and element.text.strip():
                    text_parts.append(element.text.strip())
                
                for child in element:
                    extract_text(child)
                
                if element.tail and element.tail.strip():
                    text_parts.append(element.tail.strip())
            
            extract_text(root)
            return ' '.join(text_parts)
        except ET.ParseError:
            # If XML parsing fails, try to extract as plain text
            return self._extract_txt_text(file_path)
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF file."""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip()
        except Exception as e:
            raise ValueError(f"Error extracting PDF text: {e}")
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extract text from DOCX file."""
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            raise ValueError(f"Error extracting DOCX text: {e}")
    
    def _extract_xlsx_text(self, file_path: str) -> str:
        """Extract text from XLSX file."""
        try:
            workbook = load_workbook(file_path, data_only=True)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text.strip()
        except Exception as e:
            raise ValueError(f"Error extracting XLSX text: {e}")
    
    def _extract_txt_text(self, file_path: str) -> str:
        """Extract text from plain text file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read()
            except Exception as e:
                raise ValueError(f"Error extracting text: {e}")
    
    def _extract_csv_text(self, file_path: str) -> str:
        """Extract text from CSV file."""
        try:
            text = ""
            with open(file_path, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    text += " | ".join(row) + "\n"
            return text.strip()
        except Exception as e:
            raise ValueError(f"Error extracting CSV text: {e}")
    
    def _extract_json_text(self, file_path: str) -> str:
        """Extract text from JSON file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return json.dumps(data, indent=2)
        except Exception as e:
            raise ValueError(f"Error extracting JSON text: {e}")
    
    def _extract_md_text(self, file_path: str) -> str:
        """Extract text from Markdown file."""
        return self._extract_txt_text(file_path)
    
    def _extract_html_text(self, file_path: str) -> str:
        """Extract text from HTML file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                # Simple HTML tag removal
                import re
                text = re.sub(r'<[^>]+>', '', content)
                text = re.sub(r'\s+', ' ', text)
                return text.strip()
        except Exception as e:
            raise ValueError(f"Error extracting HTML text: {e}")
    
    def count_tokens(self, text: str) -> int:
        """Count tokens in text using tiktoken."""
        return len(self.tokenizer.encode(text))
    
    def chunk_text(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """Split text into overlapping chunks."""
        tokens = self.tokenizer.encode(text)
        chunks = []
        
        for i in range(0, len(tokens), chunk_size - overlap):
            chunk_tokens = tokens[i:i + chunk_size]
            chunk_text = self.tokenizer.decode(chunk_tokens)
            chunks.append(chunk_text)
        
        return chunks
